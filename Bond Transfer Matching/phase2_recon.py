"""
phase2_recon.py
---------------
Phase 2 core logic (no Streamlit UI), including KB HUF date-range filter.
"""

import io
from typing import Optional, Tuple, Union, Dict

import pandas as pd
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process


def _normalize(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
        return ""
    return str(x).strip()


def _up(x) -> str:
    return _normalize(x).upper()


def _to_date(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce", dayfirst=True)


def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s.astype(str).str.replace(",", "", regex=False).str.strip(), errors="coerce")


def load_macro_sheets(macro_file: Union[bytes, object]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    macro_bytes = macro_file if isinstance(macro_file, (bytes, bytearray)) else macro_file.getvalue()
    bio = io.BytesIO(macro_bytes)

    bond = pd.read_excel(bio, sheet_name="Bond Info", header=0, dtype=str, engine="openpyxl")
    bio.seek(0)
    kbhuf = pd.read_excel(bio, sheet_name="KB HUF", header=0, dtype=str, engine="openpyxl")

    bond2 = bond.iloc[:, :2].copy()
    bond2.columns = ["BondName", "ISIN"]
    bond2["ISIN"] = bond2["ISIN"].apply(_up)
    bond2["BondName"] = bond2["BondName"].apply(_normalize)

    if kbhuf.shape[1] < 12:
        raise ValueError("KB HUF sheet has fewer than 12 columns; expected at least up to Col L.")

    kb2 = kbhuf.iloc[:, [1, 3, 4, 7, 11]].copy()
    kb2.columns = ["KB_ISIN", "KB_ClientName", "KB_Units", "KB_Date", "KB_Status"]

    kb2["KB_ISIN"] = kb2["KB_ISIN"].apply(_up)
    kb2["KB_ClientName"] = kb2["KB_ClientName"].apply(_normalize)
    kb2["KB_Units_num"] = _to_num(kb2["KB_Units"])
    kb2["KB_Date_dt"] = _to_date(kb2["KB_Date"])
    kb2["KB_Status_u"] = kb2["KB_Status"].apply(_up)
    kb2["KB_ClientName_key"] = kb2["KB_ClientName"].str.upper().str.strip()

    return bond2, kb2


def _build_kb_index(kb: pd.DataFrame) -> Dict[Tuple[str, float], list]:
    index: Dict[Tuple[str, float], list] = {}
    for i, r in kb.iterrows():
        key = (r["KB_ISIN"], r["KB_Units_num"])
        index.setdefault(key, []).append(i)
    return index


def _best_fuzzy_kb_row(phase_name: str, kb_candidates: pd.DataFrame, threshold: int = 95):
    if kb_candidates.empty:
        return None, None, None

    phase_key = phase_name.upper().strip()
    names = kb_candidates["KB_ClientName_key"].tolist()

    match = process.extractOne(phase_key, names, scorer=fuzz.ratio)
    if not match:
        return None, None, None

    matched_name, score, idx_in_list = match
    if score < threshold:
        return None, score, matched_name

    best_row = kb_candidates.iloc[idx_in_list]
    return best_row.name, score, best_row["KB_ClientName"]


def reconcile_phase2_fuzzy(
    phase1_df: pd.DataFrame,
    bondinfo: pd.DataFrame,
    kb: pd.DataFrame,
    name_threshold: int = 95,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> pd.DataFrame:
    p1 = phase1_df.copy()

    kb2 = kb.copy()
    if kb_date_from is not None:
        kb2 = kb2[kb2["KB_Date_dt"].notna() & (kb2["KB_Date_dt"] >= kb_date_from)]
    if kb_date_to is not None:
        kb2 = kb2[kb2["KB_Date_dt"].notna() & (kb2["KB_Date_dt"] <= kb_date_to)]

    isin_to_bond = (
        bondinfo.dropna(subset=["ISIN"])
        .drop_duplicates("ISIN")
        .set_index("ISIN")["BondName"]
        .to_dict()
    )
    p1["BondName"] = p1["ISIN"].map(isin_to_bond).fillna("")

    kb_index = _build_kb_index(kb2)

    rows = []
    for _, r in p1.iterrows():
        key = (r["ISIN"], r["Units_num"])
        idxs = kb_index.get(key, [])
        kb_candidates = kb2.loc[idxs] if idxs else kb2.iloc[0:0]

        best_idx, score, matched_name = _best_fuzzy_kb_row(r["ClientName"], kb_candidates, threshold=name_threshold)

        if best_idx is None:
            rows.append({
                "PostedDate": r["PostedDate"], "ISIN": r["ISIN"], "BondName": r.get("BondName",""),
                "Units_num": r["Units_num"], "ClientName": r["ClientName"],
                "KB_ClientName": matched_name or "", "NameMatchScore": score if score is not None else "",
                "KB_Date_dt": pd.NaT, "KB_Status": "",
                "ReconStatus": "NO_MATCH_IN_KB", "FlagColor": "Red",
                "Reason": "No KB match for same ISIN+Units with ClientName >= threshold (after KB date filter)"
            })
            continue

        kbrow = kb2.loc[best_idx]
        pdate = r["PostedDate"]
        kbdate = kbrow["KB_Date_dt"]

        if pd.isna(pdate) or pd.isna(kbdate):
            date_flag = "DATE_MISSING_REVIEW"
            date_note = "PostedDate or KB date missing"
        else:
            if pdate.date() == kbdate.date():
                date_flag = "DATE_MATCH"
                date_note = "Exact date match"
            elif pdate >= kbdate:
                date_flag = "DATE_AFTER_REVIEW"
                date_note = "PostedDate after KB date (review)"
            else:
                date_flag = "DATE_BEFORE_MISMATCH"
                date_note = "PostedDate before KB date (mismatch)"

        is_transferred = (kbrow["KB_Status_u"] == "TRANSFERRED")

        if date_flag == "DATE_MATCH" and is_transferred:
            final, color, reason = "OK", "Green", "ISIN+Units matched; Name fuzzy matched; Date matched; Status=Transferred"
        elif date_flag == "DATE_AFTER_REVIEW" and is_transferred:
            final, color, reason = "REVIEW_DATE_AFTER", "Yellow", "ISIN+Units matched; Name fuzzy matched; Status=Transferred; Date differs but PostedDate>=KB date"
        elif date_flag in ("DATE_MATCH", "DATE_AFTER_REVIEW") and not is_transferred:
            final, color, reason = "STATUS_NOT_TRANSFERRED", "Red", f"{date_note}; Status not Transferred"
        elif date_flag == "DATE_BEFORE_MISMATCH":
            final, color, reason = "DATE_BEFORE_MISMATCH", "Red", date_note
        else:
            final, color, reason = "MISSING_DATE_REVIEW", "Grey", date_note

        rows.append({
            "PostedDate": r["PostedDate"], "ISIN": r["ISIN"], "BondName": r.get("BondName",""),
            "Units_num": r["Units_num"], "ClientName": r["ClientName"],
            "KB_ClientName": kbrow["KB_ClientName"], "NameMatchScore": score,
            "KB_Date_dt": kbdate, "KB_Status": kbrow["KB_Status"],
            "ReconStatus": final, "FlagColor": color, "Reason": reason
        })

    out = pd.DataFrame(rows)
    return out.sort_values(["ReconStatus", "PostedDate", "ISIN"], na_position="last")


def run_phase2(
    phase1_df: pd.DataFrame,
    macro_file: Union[bytes, object],
    name_threshold: int = 95,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    bondinfo, kbhuf = load_macro_sheets(macro_file)
    recon = reconcile_phase2_fuzzy(
        phase1_df=phase1_df,
        bondinfo=bondinfo,
        kb=kbhuf,
        name_threshold=name_threshold,
        kb_date_from=kb_date_from,
        kb_date_to=kb_date_to,
    )
    exceptions = recon[recon["ReconStatus"] != "OK"].copy()
    return recon, exceptions


def to_colored_excel_bytes(
    phase1_df: pd.DataFrame,
    recon: pd.DataFrame,
    exceptions: pd.DataFrame,
    include_phase1_sheet: bool = True,
) -> bytes:
    buf = io.BytesIO()
    fills = {
        "Green": PatternFill("solid", fgColor="C6EFCE"),
        "Yellow": PatternFill("solid", fgColor="FFEB9C"),
        "Red": PatternFill("solid", fgColor="FFC7CE"),
        "Grey": PatternFill("solid", fgColor="D9D9D9"),
    }

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if include_phase1_sheet:
            phase1_df.to_excel(writer, index=False, sheet_name="Phase1_Debits")

        recon.to_excel(writer, index=False, sheet_name="Reconciliation")
        exceptions.to_excel(writer, index=False, sheet_name="Exceptions")

        wb = writer.book
        for sheet_name in ["Reconciliation", "Exceptions"]:
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            if "FlagColor" not in header:
                continue
            flag_col = header.index("FlagColor") + 1
            for row in range(2, ws.max_row + 1):
                flag = ws.cell(row=row, column=flag_col).value
                fill = fills.get(flag, None)
                if fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill

    buf.seek(0)
    return buf.getvalue()
