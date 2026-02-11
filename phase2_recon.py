"""
phase2_recon.py
---------------
Phase 2 core logic (no Streamlit UI), including:
- KB HUF date-range filter (based on KB_Date_dt / Col H)
- Fuzzy client-name matching (default threshold 95)
- Bidirectional reconciliation:
    1) DP/Phase-1 -> KB HUF (Reconciliation + Exceptions)
    2) KB HUF -> DP/Phase-1 (KB_Unmatched: entries in KB HUF within the date filter that did NOT match any DP/Phase-1 row)
"""

import io
from typing import Optional, Tuple, Union, Dict

import pandas as pd
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process


def _normalize(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
        return ""
    return str(x).strip()


def _up(x) -> str:
    return _normalize(x).upper()


def _to_date(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce", dayfirst=True)


def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s.astype(str).str.replace(",", "", regex=False).str.strip(), errors="coerce")


def load_macro_sheets(macro_file: Union[bytes, object]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Reads only:
      - Bond Info: Col A BondName, Col B ISIN
      - KB HUF:    Col B ISIN, Col D ClientName, Col E Units, Col H Date, Col L Status
    """
    macro_bytes = macro_file if isinstance(macro_file, (bytes, bytearray)) else macro_file.getvalue()
    bio = io.BytesIO(macro_bytes)

    bond = pd.read_excel(bio, sheet_name="Bond Info", header=0, dtype=str, engine="openpyxl")
    bio.seek(0)
    kbhuf = pd.read_excel(bio, sheet_name="KB HUF", header=0, dtype=str, engine="openpyxl")

    # Bond Info: A=BondName, B=ISIN
    bond2 = bond.iloc[:, :2].copy()
    bond2.columns = ["BondName", "ISIN"]
    bond2["ISIN"] = bond2["ISIN"].apply(_up)
    bond2["BondName"] = bond2["BondName"].apply(_normalize)

    # KB HUF: B, D, E, H, L by position
    if kbhuf.shape[1] < 12:
        raise ValueError("KB HUF sheet has fewer than 12 columns; expected at least up to Col L.")

    kb2 = kbhuf.iloc[:, [1, 3, 4, 7, 11]].copy()
    kb2.columns = ["KB_ISIN", "KB_ClientName", "KB_Units", "KB_Date", "KB_Status"]

    kb2["KB_ISIN"] = kb2["KB_ISIN"].apply(_up)
    kb2["KB_ClientName"] = kb2["KB_ClientName"].apply(_normalize)
    kb2["KB_Units_num"] = _to_num(kb2["KB_Units"])
    kb2["KB_Date_dt"] = _to_date(kb2["KB_Date"])
    kb2["KB_Status_u"] = kb2["KB_Status"].apply(_up)
    kb2["KB_ClientName_key"] = kb2["KB_ClientName"].str.upper().str.strip()

    return bond2, kb2


def _build_kb_index(kb: pd.DataFrame) -> Dict[Tuple[str, float], list]:
    """
    Index by strict keys:
      (ISIN, Units_num) -> list[row_index]
    """
    index: Dict[Tuple[str, float], list] = {}
    for i, r in kb.iterrows():
        key = (r["KB_ISIN"], r["KB_Units_num"])
        index.setdefault(key, []).append(i)
    return index


def _best_fuzzy_kb_row(phase_name: str, kb_candidates: pd.DataFrame, threshold: int = 95):
    """
    Return (best_row_index, best_score, matched_name) or (None, score, matched_name).
    """
    if kb_candidates.empty:
        return None, None, None

    phase_key = phase_name.upper().strip()
    names = kb_candidates["KB_ClientName_key"].tolist()

    match = process.extractOne(phase_key, names, scorer=fuzz.ratio)
    if not match:
        return None, None, None

    matched_name, score, idx_in_list = match
    if score < threshold:
        return None, score, matched_name

    best_row = kb_candidates.iloc[idx_in_list]
    return best_row.name, score, best_row["KB_ClientName"]


def _apply_kb_date_filter(
    kb: pd.DataFrame,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> pd.DataFrame:
    kb2 = kb.copy()
    if kb_date_from is not None:
        kb2 = kb2[kb2["KB_Date_dt"].notna() & (kb2["KB_Date_dt"] >= kb_date_from)]
    if kb_date_to is not None:
        kb2 = kb2[kb2["KB_Date_dt"].notna() & (kb2["KB_Date_dt"] <= kb_date_to)]
    return kb2


def reconcile_phase2_fuzzy(
    phase1_df: pd.DataFrame,
    bondinfo: pd.DataFrame,
    kb: pd.DataFrame,
    name_threshold: int = 95,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      - recon_df: DP/Phase-1 -> KB reconciliation (one row per Phase-1 txn)
      - exceptions_df: subset of recon_df where ReconStatus != OK
      - kb_unmatched_df: KB rows (within the KB date filter) that did NOT match any Phase-1 txn
    """
    p1 = phase1_df.copy()

    kb2 = _apply_kb_date_filter(kb, kb_date_from=kb_date_from, kb_date_to=kb_date_to)

    isin_to_bond = (
        bondinfo.dropna(subset=["ISIN"])
        .drop_duplicates("ISIN")
        .set_index("ISIN")["BondName"]
        .to_dict()
    )
    p1["BondName"] = p1["ISIN"].map(isin_to_bond).fillna("")

    kb_index = _build_kb_index(kb2)

    matched_kb_indices = set()

    rows = []
    for _, r in p1.iterrows():
        key = (r["ISIN"], r["Units_num"])
        idxs = kb_index.get(key, [])
        kb_candidates = kb2.loc[idxs] if idxs else kb2.iloc[0:0]

        best_idx, score, matched_name = _best_fuzzy_kb_row(
            r["ClientName"], kb_candidates, threshold=name_threshold
        )

        if best_idx is None:
            rows.append({
                "PostedDate": r["PostedDate"], "ISIN": r["ISIN"], "BondName": r.get("BondName", ""),
                "Units_num": r["Units_num"], "ClientName": r["ClientName"],
                "KB_ClientName": matched_name or "", "NameMatchScore": score if score is not None else "",
                "KB_Date_dt": pd.NaT, "KB_Status": "",
                "ReconStatus": "NO_MATCH_IN_KB", "FlagColor": "Red",
                "Reason": "No KB match for same ISIN+Units with ClientName >= threshold (after KB date filter)"
            })
            continue

        matched_kb_indices.add(best_idx)

        kbrow = kb2.loc[best_idx]
        pdate = r["PostedDate"]
        kbdate = kbrow["KB_Date_dt"]

        if pd.isna(pdate) or pd.isna(kbdate):
            date_flag = "DATE_MISSING_REVIEW"
            date_note = "PostedDate or KB date missing"
        else:
            if pdate.date() == kbdate.date():
                date_flag = "DATE_MATCH"
                date_note = "Exact date match"
            elif pdate >= kbdate:
                date_flag = "DATE_AFTER_REVIEW"
                date_note = "PostedDate after KB date (review)"
            else:
                date_flag = "DATE_BEFORE_MISMATCH"
                date_note = "PostedDate before KB date (mismatch)"

        is_transferred = (kbrow["KB_Status_u"] == "TRANSFERRED")

        if date_flag == "DATE_MATCH" and is_transferred:
            final, color, reason = "OK", "Green", "ISIN+Units matched; Name fuzzy matched; Date matched; Status=Transferred"
        elif date_flag == "DATE_AFTER_REVIEW" and is_transferred:
            final, color, reason = "REVIEW_DATE_AFTER", "Yellow", "ISIN+Units matched; Name fuzzy matched; Status=Transferred; Date differs but PostedDate>=KB date"
        elif date_flag in ("DATE_MATCH", "DATE_AFTER_REVIEW") and not is_transferred:
            final, color, reason = "STATUS_NOT_TRANSFERRED", "Red", f"{date_note}; Status not Transferred"
        elif date_flag == "DATE_BEFORE_MISMATCH":
            final, color, reason = "DATE_BEFORE_MISMATCH", "Red", date_note
        else:
            final, color, reason = "MISSING_DATE_REVIEW", "Grey", date_note

        rows.append({
            "PostedDate": r["PostedDate"], "ISIN": r["ISIN"], "BondName": r.get("BondName", ""),
            "Units_num": r["Units_num"], "ClientName": r["ClientName"],
            "KB_ClientName": kbrow["KB_ClientName"], "NameMatchScore": score,
            "KB_Date_dt": kbdate, "KB_Status": kbrow["KB_Status"],
            "ReconStatus": final, "FlagColor": color, "Reason": reason
        })

    recon = pd.DataFrame(rows).sort_values(["ReconStatus", "PostedDate", "ISIN"], na_position="last")
    exceptions = recon[recon["ReconStatus"] != "OK"].copy()

    # --- Reverse check: KB rows not matched to any DP/Phase-1 row (within KB date filter) ---
    kb_unmatched = kb2.loc[~kb2.index.isin(matched_kb_indices)].copy()
    if not kb_unmatched.empty:
        kb_unmatched["BondName"] = kb_unmatched["KB_ISIN"].map(isin_to_bond).fillna("")

        def _kb_only_flag(row):
            if row.get("KB_Status_u", "") == "TRANSFERRED":
                return "KB_ONLY_NO_DP_MATCH", "Red", "KB entry marked Transferred but no matching DP/Phase-1 transaction"
            if row.get("KB_Status_u", "") in ("TRANSFER PENDING", "PENDING", "NOT TRANSFERRED"):
                return "KB_ONLY_PENDING", "Yellow", "KB entry exists but not Transferred and no matching DP/Phase-1 transaction"
            return "KB_ONLY_REVIEW", "Grey", "KB entry exists with no DP/Phase-1 match (review)"

        tmp = kb_unmatched.apply(
            lambda r: pd.Series(_kb_only_flag(r), index=["ReconStatus", "FlagColor", "Reason"]),
            axis=1,
        )
        kb_unmatched = pd.concat([kb_unmatched, tmp], axis=1)

        kb_unmatched_df = kb_unmatched[[
            "KB_Date_dt", "KB_ISIN", "BondName", "KB_Units_num", "KB_ClientName", "KB_Status",
            "ReconStatus", "FlagColor", "Reason"
        ]].rename(columns={
            "KB_Date_dt": "KB_Date",
            "KB_ISIN": "ISIN",
            "KB_Units_num": "Units_num",
            "KB_ClientName": "ClientName",
        }).sort_values(["ReconStatus", "KB_Date", "ISIN"], na_position="last")
    else:
        kb_unmatched_df = pd.DataFrame(columns=[
            "KB_Date", "ISIN", "BondName", "Units_num", "ClientName", "KB_Status",
            "ReconStatus", "FlagColor", "Reason"
        ])

    return recon, exceptions, kb_unmatched_df


def run_phase2(
    phase1_df: pd.DataFrame,
    macro_file: Union[bytes, object],
    name_threshold: int = 95,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Main entrypoint for Phase-2.
    Returns: recon_df, exceptions_df, kb_unmatched_df
    """
    bondinfo, kbhuf = load_macro_sheets(macro_file)
    return reconcile_phase2_fuzzy(
        phase1_df=phase1_df,
        bondinfo=bondinfo,
        kb=kbhuf,
        name_threshold=name_threshold,
        kb_date_from=kb_date_from,
        kb_date_to=kb_date_to,
    )


def run_phase2_legacy(
    phase1_df: pd.DataFrame,
    macro_file: Union[bytes, object],
    name_threshold: int = 95,
    kb_date_from: Optional[pd.Timestamp] = None,
    kb_date_to: Optional[pd.Timestamp] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Backwards-compatible wrapper if older orchestrator expects only recon + exceptions."""
    recon, exceptions, _kb_unmatched = run_phase2(
        phase1_df=phase1_df,
        macro_file=macro_file,
        name_threshold=name_threshold,
        kb_date_from=kb_date_from,
        kb_date_to=kb_date_to,
    )
    return recon, exceptions


def to_colored_excel_bytes(
    phase1_df: pd.DataFrame,
    recon: pd.DataFrame,
    exceptions: pd.DataFrame,
    kb_unmatched: Optional[pd.DataFrame] = None,
    include_phase1_sheet: bool = True,
    include_kb_unmatched_sheet: bool = True,
) -> bytes:
    """
    Writes a single workbook with:
      - Phase1_Debits (optional)
      - Reconciliation (colored by FlagColor)
      - Exceptions (colored by FlagColor)
      - KB_Unmatched (optional; colored by FlagColor)
    """
    buf = io.BytesIO()
    fills = {
        "Green": PatternFill("solid", fgColor="C6EFCE"),
        "Yellow": PatternFill("solid", fgColor="FFEB9C"),
        "Red": PatternFill("solid", fgColor="FFC7CE"),
        "Grey": PatternFill("solid", fgColor="D9D9D9"),
    }

    kb_unmatched = kb_unmatched if kb_unmatched is not None else pd.DataFrame()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if include_phase1_sheet:
            phase1_df.to_excel(writer, index=False, sheet_name="Phase1_Debits")

        recon.to_excel(writer, index=False, sheet_name="Reconciliation")
        exceptions.to_excel(writer, index=False, sheet_name="Exceptions")

        if include_kb_unmatched_sheet and not kb_unmatched.empty:
            kb_unmatched.to_excel(writer, index=False, sheet_name="KB_Unmatched")

        wb = writer.book

        def _color_sheet(sheet_name: str):
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            if "FlagColor" not in header:
                return
            flag_col = header.index("FlagColor") + 1
            for row in range(2, ws.max_row + 1):
                flag = ws.cell(row=row, column=flag_col).value
                fill = fills.get(flag, None)
                if fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill

        for sheet_name in ["Reconciliation", "Exceptions"]:
            _color_sheet(sheet_name)

        if include_kb_unmatched_sheet and not kb_unmatched.empty:
            _color_sheet("KB_Unmatched")

    buf.seek(0)
    return buf.getvalue()
